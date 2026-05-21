from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.application import Application


TRACKER_COLUMNS = [
    "Date Applied",
    "Company",
    "Job Title",
    "Job URL",
    "Source Website",
    "Location",
    "Remote/Hybrid/Onsite",
    "Salary Range",
    "Resume Version",
    "Cover Letter Version",
    "Application Status",
    "Submitted By Extension",
    "Follow-up Date",
    "Recruiter Contact",
    "Notes",
    "Last Updated",
]

STATUS_VALUES = ["draft", "in_progress", "needs_user_action", "ready_for_review", "submitted", "abandoned"]


def _set_column_widths(ws: Worksheet) -> None:
    widths = {
        "A": 13,
        "B": 22,
        "C": 22,
        "D": 45,
        "E": 16,
        "F": 18,
        "G": 20,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 18,
        "L": 20,
        "M": 14,
        "N": 20,
        "O": 36,
        "P": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def export_applications_xlsx(applications: Iterable[Application]) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    ws.append(TRACKER_COLUMNS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P1"
    _set_column_widths(ws)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Status dropdown
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_VALUES)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("K2:K1048576")

    for app in applications:
        review = app.review_packet or {}
        job_meta = review.get("job_meta") or {}
        salary_range = job_meta.get("salary_range") or ""
        remote = job_meta.get("workplace_type") or ""

        date_applied = app.submitted_at.date().isoformat() if app.submitted_at else ""
        last_updated = (app.updated_at or app.created_at).astimezone(timezone.utc).isoformat()

        ws.append(
            [
                date_applied,
                job_meta.get("company") or "",
                job_meta.get("job_title") or "",
                job_meta.get("job_url") or "",
                app.application_source,
                job_meta.get("location") or "",
                remote,
                salary_range,
                app.resume_version or "",
                app.cover_letter_version or "",
                app.status,
                "yes" if app.submitted_by_extension else "no",
                app.follow_up_date.date().isoformat() if app.follow_up_date else "",
                app.recruiter_contact or "",
                app.notes or "",
                last_updated,
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"job_applications_{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return buf.read(), filename


def import_applications_xlsx(xlsx_bytes: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    col_index = {name: idx for idx, name in enumerate(header)}

    def get(row, col_name: str) -> str:
        idx = col_index.get(col_name)
        if idx is None:
            return ""
        val = row[idx]
        if val is None:
            return ""
        if isinstance(val, datetime):
            return val.date().isoformat()
        return str(val)

    imported: list[dict] = []
    for row in rows[1:]:
        if row is None:
            continue
        job_url = get(row, "Job URL")
        if not job_url:
            continue
        imported.append(
            {
                "date_applied": get(row, "Date Applied"),
                "company": get(row, "Company"),
                "job_title": get(row, "Job Title"),
                "job_url": job_url,
                "source_website": get(row, "Source Website") or "other",
                "location": get(row, "Location"),
                "workplace_type": get(row, "Remote/Hybrid/Onsite"),
                "salary_range": get(row, "Salary Range"),
                "resume_version": get(row, "Resume Version") or None,
                "cover_letter_version": get(row, "Cover Letter Version") or None,
                "status": get(row, "Application Status") or "draft",
                "submitted_by_extension": (get(row, "Submitted By Extension").lower() in ("yes", "true", "1")),
                "follow_up_date": get(row, "Follow-up Date") or None,
                "recruiter_contact": get(row, "Recruiter Contact") or None,
                "notes": get(row, "Notes") or None,
            }
        )
    return imported

